from pathlib import Path
import openpyxl

from .base import BrokerImporter, ParsedAccount, ParsedHolding, ParseResult
from .etf_tickers import LEVERAGED_ETF_TICKERS, KNOWN_ETF_TICKERS

# Questrade account-type label → our schema enum
ACCOUNT_TYPE_MAP = {
    "Individual margin": "Unreg",
    "Individual cash":   "Unreg",
    "Individual":        "Unreg",
    "Individual TFSA":   "TFSA",
    "Individual RRSP":   "RRSP",
    "Individual LIRA":   "RRSP",   # treat LIRA as RRSP-equivalent for v0.1
    "Individual RESP":   "Unreg",  # not in enum; flag instead if encountered
}

# Questrade Asset Class code → our enum
ASSET_CLASS_MAP = {
    "STK":   "Stock",
    "ETF":   "ETF",
    "CASH":  "Cash",
    "FUND":  "ETF",
    "BOND":  "Stock",   # no bond enum in v0.1; classify as Stock
    "OPT":   "Stock",
}


def _refine_asset_class(ticker: str, raw_code: str, description: str) -> str:
    t = ticker.upper()
    if t in LEVERAGED_ETF_TICKERS:
        return "LeveragedETF"
    base = ASSET_CLASS_MAP.get((raw_code or "").upper(), "Stock")
    return base


def _default_country(ticker: str, currency: str, description: str) -> str:
    if currency == "CAD":
        return "CA"
    if currency == "USD":
        return "US"
    return "Other"


# Yahoo doesn't use the same symbol Questrade displays for a few TSX-listed
# USD-denominated ETFs. Hand-override them here. Add as we find more.
YAHOO_SYMBOL_OVERRIDES = {
    "HSUV.U": "HSUV-U.TO",
    "UCSH.U": "UCSH-U.TO",
    "ETHH.B": "ETHH-B.TO",
}


def _to_yahoo_ticker(ticker: str, currency: str, description: str) -> str:
    """Map Questrade display ticker to a Yahoo Finance symbol."""
    t = (ticker or "").strip()
    if not t:
        return t
    if t in YAHOO_SYMBOL_OVERRIDES:
        return YAHOO_SYMBOL_OVERRIDES[t]
    if currency == "USD":
        return t
    desc = (description or "").upper()
    if "CDR" in desc:
        base = t.split(".")[0]
        return f"{base}.NE"
    if t.endswith((".TO", ".V", ".NE", ".CN")):
        return t
    # Yahoo uses dash for class/unit suffixes on TSX listings:
    # AP.UN -> AP-UN.TO, HR.UN -> HR-UN.TO, ETHH.B -> ETHH-B.TO (also in overrides).
    return f"{t.replace('.', '-')}.TO"


class QuestradeImporter(BrokerImporter):
    broker = "Questrade"

    @staticmethod
    def detect_format(path: Path) -> bool:
        if path.suffix.lower() != ".xlsx":
            return False
        stem = path.stem.lower()
        return "questrade" in stem or "investmentsummary" in stem.replace(" ", "").replace(".", "")

    def parse(self, path: Path) -> ParseResult:
        wb = openpyxl.load_workbook(path, data_only=True)
        result = ParseResult()

        balances = wb["Balances"]
        headers = [c.value for c in balances[1]]
        ix = {h: i for i, h in enumerate(headers) if h}
        for row in balances.iter_rows(min_row=2, values_only=True):
            if row[ix["Account Number"]] is None:
                continue
            num = str(row[ix["Account Number"]]).strip()
            raw_type = str(row[ix["Account Type"]]).strip()
            mapped = ACCOUNT_TYPE_MAP.get(raw_type)
            if mapped is None:
                raise ValueError(
                    f"Unknown Questrade account type {raw_type!r}; add it to ACCOUNT_TYPE_MAP"
                )
            cash = float(row[ix["Cash in CAD Combined"]] or 0)
            result.accounts[num] = ParsedAccount(
                broker="Questrade",
                account_type=mapped,
                label=f"{raw_type} — {num}",
                cash_cad=cash,
                broker_account_number=num,
            )

        positions = wb["Positions"]
        pheaders = [c.value for c in positions[1]]
        px = {h: i for i, h in enumerate(pheaders) if h}
        for row in positions.iter_rows(min_row=2, values_only=True):
            sym = row[px["Equity Symbol"]]
            if sym is None:
                continue
            desc = row[px["Equity Description"]] or ""
            acct = str(row[px["Account Number"]]).strip()
            cur = (row[px["Currency"]] or "USD").upper()
            asset_code = row[px["Asset Class"]] or ""
            qty = float(row[px["Quantity"]])
            cps = float(row[px["Cost Per Share"]])
            ac = _refine_asset_class(str(sym), asset_code, desc)
            country = _default_country(str(sym), cur, desc)
            yahoo = _to_yahoo_ticker(str(sym), cur, desc)
            result.holdings.append(ParsedHolding(
                broker_account_number=acct,
                ticker=str(sym),
                yahoo_ticker=yahoo,
                currency=cur,
                quantity=qty,
                acb_per_share=cps,
                asset_class=ac,
                country=country,
                description=str(desc),
            ))

        return result
